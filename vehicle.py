import numpy as np
import cv2
import datetime
import os
import openpyxl
import pytesseract
import re

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
count = 0
mainFilename = f"{datetime.datetime.now().day}.{datetime.datetime.now().month}.{datetime.datetime.now().year}.xlsx"
cascade_path = 'licenceplate.xml'

if not os.path.isfile(cascade_path):
    print(f"Error: Cascade file '{cascade_path}' not found.")
    exit()
plate_cascade = cv2.CascadeClassifier(cascade_path)

if plate_cascade.empty():
    print("Error: Failed to load cascade classifier.")
    exit()

video_path = 'video2.mov'
cap = cv2.VideoCapture(video_path)

if not cap.isOpened():
    print(f"Error: Could not open video file '{video_path}'.")
    exit()

if not os.path.isfile(mainFilename):
    mwb = openpyxl.Workbook()
    main_sheet = mwb.active
    main_sheet.cell(row=1, column=1).value = 'Date'
    main_sheet.cell(row=1, column=2).value = 'Time'
    main_sheet.cell(row=1, column=3).value = 'Vehicle_Number'
    mwb.save(mainFilename)
    print(f"Created new Excel file '{mainFilename}'")
else:
    print(f"Using existing Excel file '{mainFilename}'")

def get_string(plate_img):
    gray = cv2.cvtColor(plate_img, cv2.COLOR_BGR2GRAY)
    kernel = np.ones((1, 1), np.uint8)
    img = cv2.dilate(gray, kernel, iterations=1)
    img = cv2.erode(img, kernel, iterations=1)
    img = cv2.GaussianBlur(img, (5, 5), 0)  
    result = pytesseract.image_to_string(img, config='--psm 8') 
    return result.strip()

def correct_plate_number(plate_number):
    corrections = {
        'LN': 'TN',
        'lN': 'TN',
        'lM': 'TN',
        'IN': 'TN',
       
    }
    for wrong, correct in corrections.items():
        if wrong in plate_number:
            plate_number = plate_number.replace(wrong, correct)
    return plate_number

def is_valid_plate(plate_number):
    pattern = r'^[A-Z]{2}[0-9]{1,2}[A-Z]{1,2}[0-9]{4}$'
    return re.match(pattern, plate_number) is not None

def clean_plate_number(plate_number):
    plate_number = re.sub(r'\W+', '', plate_number)  
    return plate_number

def is_plate_stored(plate_number):
    mwb = openpyxl.load_workbook(mainFilename)
    main_sheet = mwb.active
    for row in main_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        if row[2] and plate_number in row[2]:
            return True
    return False

def process_roi(count, img, roi, detected_plates):
    x, y, w, h = roi

    if y + 10 >= y + h - 10 or x + 10 >= x + w - 10 or y + h - 10 > img.shape[0] or x + w - 10 > img.shape[1]:
        print(f"Skipping invalid ROI: {roi}")
        return
    
    roi_plate = img[y + 10:y + h - 10, x + 10:x + w - 10]
    plate_img = roi_plate.copy()

    if plate_img.size == 0:
        print(f"Empty plate image at ROI: {roi}")
        return

    plate_number = get_string(plate_img)
    plate_number = clean_plate_number(plate_number)
    plate_number = correct_plate_number(plate_number)

    if is_valid_plate(plate_number) and plate_number not in detected_plates:
        if not is_plate_stored(plate_number):
            print(f"Valid Detected Number {count}: {plate_number}")
            detected_plates.add(plate_number)
            log_to_excel(plate_number)
        else:
            print(f"Plate Number {count} already stored: {plate_number}")
    else:
        print(f"Invalid or Duplicate Plate Number {count}: {plate_number}")

def log_to_excel(plate_number):
    now = datetime.datetime.now()
    current_date = f"{now.day:02}/{now.month:02}/{now.year}"
    current_time = f"{now.hour:02}:{now.minute:02}:{now.second:02}"
    filename = mainFilename

    mwb = openpyxl.load_workbook(filename)
    main_sheet = mwb.active
    max_row = main_sheet.max_row
    main_sheet.cell(row=max_row + 1, column=1).value = current_date
    main_sheet.cell(row=max_row + 1, column=2).value = current_time
    main_sheet.cell(row=max_row + 1, column=3).value = plate_number
    mwb.save(filename)
    print(f"Logged to Excel: {plate_number} at {current_time} on {current_date}")

detected_plates = set()
while True:
    ret, img = cap.read()
    if not ret:
        print("Error: Could not read frame from video.")
        break

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    plates = plate_cascade.detectMultiScale(gray, 1.3, 5)

    for (x, y, w, h) in plates:
        count += 1
        roi = (x, y, w, h)
        process_roi(count, img, roi, detected_plates)

    cv2.imshow('img', img)
    if cv2.waitKey(1) & 0xFF == 27:
        break

cap.release()
cv2.destroyAllWindows()
